import io
import zipfile

from django.core.exceptions import ObjectDoesNotExist
from django.http import Http404
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from candidates_app.models import DEGREE, CandidateModel
from evaluations_app.models import CandidateEvaluationModel


def write_candidate_evaluator_data(work_sheet, candidate, evaluator):
    # candidate general info
    work_sheet['B4'] = candidate.first_name
    work_sheet['C4'] = candidate.last_name
    work_sheet['D4'] = candidate.candidate_id
    work_sheet['F4'] = DEGREE[candidate.applying_degree][1]
    # evaluator general info
    work_sheet['B5'] = evaluator.first_name
    work_sheet['C5'] = evaluator.last_name
    work_sheet['D5'] = str(evaluator.staff_id)


def get_evaluation_data(evaluation_id=None,
                        evaluation_type='application_evaluation'):
    try:
        return CandidateEvaluationModel.objects. \
            select_related('candidate',
                           'evaluator',
                           evaluation_type
                           ).get(evaluation_id=evaluation_id)
    except ObjectDoesNotExist:
        raise Http404("No {0} found matching the query".format(
            CandidateEvaluationModel._meta.verbose_name
        ))


def generate_application_xlsx(evaluation):
    candidate = evaluation.candidate
    evaluator = evaluation.evaluator
    application_evaluation = evaluation.application_evaluation
    name = "application_{0}_{1}({2}_{3}).xls".format(
        evaluation.candidate.first_name, evaluation.candidate.last_name,
        evaluator.first_name, evaluator.last_name)
    # open excel template
    url = 'static/xlsx/application_template.xlsx'
    wb = load_workbook(url)
    sheets = wb.sheetnames
    work_sheet = wb[sheets[0]]
    write_candidate_evaluator_data(work_sheet, candidate, evaluator)
    # candidate education information
    for instance in candidate.education_info.all():
        row_list = [instance.start_date, instance.end_date,
                    instance.grad_date, instance.degree_type,
                    instance.institution,
                    instance.study_field, instance.gpa]
        row_num = 8
        for col_num, value in enumerate(row_list):
            work_sheet.cell(row=row_num, column=col_num + 1).value = value
        row_num += 1
    # candidate testing information
    work_sheet['B12'] = candidate.testing_info.ielts
    work_sheet['E12'] = candidate.testing_info.gre
    work_sheet['B13'] = candidate.testing_info.toefl
    # evaluated by secretary
    work_sheet['F16'] = candidate.gpa
    work_sheet['F17'] = candidate.school_rating
    work_sheet['F18'] = candidate.research_experience
    # evaluated by evaluator
    work_sheet['F19'] = application_evaluation.relevancy
    work_sheet['F20'] = application_evaluation.statement_of_purpose
    work_sheet['F21'] = application_evaluation.recommendation_1
    work_sheet['F22'] = application_evaluation.recommendation_2
    work_sheet['F23'] = application_evaluation.relevant_degrees
    work_sheet['B24'] = application_evaluation.evaluation_comment
    return [wb, name]


def generate_interview_xlsx(evaluation):
    candidate = evaluation.candidate
    evaluator = evaluation.evaluator
    interview_evaluation = evaluation.interview_evaluation
    name = "interview_{0}_{1}({2}_{3}).xls".format(
        candidate.first_name, candidate.last_name,
        evaluator.first_name, evaluator.last_name)
    # open excel template
    url = 'static/xlsx/interview_template.xlsx'
    wb = load_workbook(url)
    sheets = wb.sheetnames
    work_sheet = wb[sheets[0]]
    write_candidate_evaluator_data(work_sheet, candidate, evaluator)
    # evaluated by evaluator
    work_sheet['F8'] = interview_evaluation.work_experience_goals
    work_sheet['F9'] = interview_evaluation. \
        research_interest_and_motivation
    work_sheet['F10'] = interview_evaluation.understanding_of_major
    work_sheet['F11'] = interview_evaluation.community_involvement
    work_sheet['F12'] = interview_evaluation.interpersonal_skills
    work_sheet['F13'] = interview_evaluation.english_level
    work_sheet['B14'] = interview_evaluation.interview_comment
    return [wb, name]


def save_application_zip(candidate_id):
    candidate = CandidateModel.objects.get(candidate_id=candidate_id)
    evaluations = CandidateEvaluationModel.objects.filter(candidate=candidate)
    buffer = io.BytesIO()
    file_name = "application_evaluations_{0}_{1}.zip".format(
        candidate.first_name,
        candidate.last_name
    )
    with zipfile.ZipFile(buffer, 'w') as zipped_f:
        for evaluation in evaluations:
            [wb, name] = generate_application_xlsx(evaluation)
            zipped_f.writestr(name, save_virtual_workbook(wb))
        zipped_f.close()
    return [buffer, file_name]


def save_interview_zip(candidate_id):
    candidate = CandidateModel.objects.get(candidate_id=candidate_id)
    evaluations = CandidateEvaluationModel.objects.filter(candidate=candidate)
    buffer = io.BytesIO()
    file_name = "interview_evaluations_{0}_{1}.zip".format(
        candidate.first_name,
        candidate.last_name
    )
    with zipfile.ZipFile(buffer, 'w') as zipped_f:
        for evaluation in evaluations:
            [wb, name] = generate_interview_xlsx(evaluation)
            zipped_f.writestr(name, save_virtual_workbook(wb))
        zipped_f.close()
    return [buffer, file_name]

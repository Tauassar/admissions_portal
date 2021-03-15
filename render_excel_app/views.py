from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook
from django.http import HttpResponse

from auth_app.decorators import check_permissions
from auth_app.models import CustomUserModel
from evaluations_app.models import CandidateEvaluationModel
from render_excel_app.utils import (write_candidate_evaluator_data,
                                    get_evaluation_data)


@login_required(login_url='login')
@check_permissions(allowed_pos=[CustomUserModel.SECRETARY])
def ApplicationEvalAsExcelView(request, evaluation_id):
    """
        Render application evaluation into excel and send back to client
    """
    evaluation = get_evaluation_data(evaluation_id,
                                     'application_evaluation')
    candidate = evaluation.candidate
    evaluator = evaluation.evaluator
    application_evaluation = evaluation.application_evaluation
    # set response type to excel file
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = \
        'attachment; filename=' \
        '"application_evaluation_{0}_{1}.xls"'.format(
            candidate.first_name, candidate.last_name)
    # open excel template
    url = 'static/xlsx/application_template.xlsx'
    wb = load_workbook(url)
    sheets = wb.sheetnames
    work_sheet = wb[sheets[0]]
    write_candidate_evaluator_data(work_sheet, candidate, evaluator)
    # candidate education information
    for instance in candidate.education_info.all():
        row_list = [instance.start_date, instance.end_date,
                    instance.grad_date, instance.degree_type,
                    instance.institution,
                    instance.study_field, instance.gpa]
        row_num = 8
        for col_num, value in enumerate(row_list):
            work_sheet.cell(row=row_num, column=col_num + 1).value = value
        row_num += 1
    # candidate testing information
    work_sheet['B12'] = candidate.testing_info.ielts
    work_sheet['E12'] = candidate.testing_info.gre
    work_sheet['B13'] = candidate.testing_info.toefl
    # evaluated by secretary
    work_sheet['F16'] = candidate.gpa
    work_sheet['F17'] = candidate.school_rating
    work_sheet['F18'] = candidate.research_experience
    # evaluated by evaluator
    work_sheet['F19'] = application_evaluation.relevancy
    work_sheet['F20'] = application_evaluation.statement_of_purpose
    work_sheet['F21'] = application_evaluation.recommendation_1
    work_sheet['F22'] = application_evaluation.recommendation_2
    work_sheet['F23'] = application_evaluation.relevant_degrees
    work_sheet['B24'] = application_evaluation.evaluation_comment
    wb.save(response)
    return response


@login_required(login_url='login')
@check_permissions(allowed_pos=[CustomUserModel.SECRETARY])
def InterviewEvalAsExcelView(request, evaluation_id):
    """
        Render interview evaluation into excel and send back to client
    """
    evaluation = get_evaluation_data(evaluation_id,
                                     'interview_evaluation')
    candidate = evaluation.candidate
    evaluator = evaluation.evaluator
    interview_evaluation = evaluation.interview_evaluation
    # set response type to excel file
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = \
        'attachment; filename="interview_evaluation_{0}_{1}.xls"' .format(
            candidate.first_name, candidate.last_name)
    # open excel template
    url = 'static/xlsx/interview_template.xlsx'
    wb = load_workbook(url)
    sheets = wb.sheetnames
    work_sheet = wb[sheets[0]]
    write_candidate_evaluator_data(work_sheet, candidate, evaluator)
    # evaluated by evaluator
    work_sheet['F8'] = interview_evaluation.work_experience_goals
    work_sheet['F9'] = interview_evaluation. \
        research_interest_and_motivation
    work_sheet['F10'] = interview_evaluation.understanding_of_major
    work_sheet['F11'] = interview_evaluation.community_involvement
    work_sheet['F12'] = interview_evaluation.interpersonal_skills
    work_sheet['F13'] = interview_evaluation.english_level
    work_sheet['B14'] = interview_evaluation.interview_comment
    wb.save(response)
    return response

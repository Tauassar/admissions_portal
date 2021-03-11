from openpyxl import load_workbook
from django.http import HttpResponse
from django.core.exceptions import ObjectDoesNotExist
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import (
    authenticate, login, logout, get_user_model, update_session_auth_hash)
from django.contrib.auth.decorators import login_required
from rest_framework import generics, permissions
from .decorators import auth_check, check_permissions
from .models import (
    CandidateModel,
    CandidateEvaluationModel,
    InformationModel,
    AdmissionYearModel,
    ApplicationEvaluationModel,
    InterviewEvaluationModel
)
from .forms import (
    CustomPasswordChangeForm,
    SecretaryEvaluationForm,
    AddCandidateForm,
    ApprovementForm,
    AdmissionRoundForm,
    ApplicationFormset,
    InterviewFormset,
    TestingFormset,
    EducationFormset)
from .serializers import (
    CandidateSerializer,
)

"""
    TODO: 
        add committee report error button
        secretary comments
        forgot pass
        mail send
        different dashboards for different users
        recent actions
        change image profile
"""
ADMISSION_DEPARTMENT = 0
COMMITTEE_MEMBER = 1
COMMITTEE_CHAIR = 2
SECRETARY = 3
POSITIONS = [
    'Admission department',
    'Admission committee member',
    'Chair of the admission committee',
    'School Secretary'
]


def get_current_year_and_round():
    """
        get active admission year and round from database
    """
    admission_year = get_object_or_404(AdmissionYearModel, active=True)
    admission_round = admission_year.get_current_admission_round()
    return [admission_year, admission_round]


@login_required(login_url='login')
def dashboardView(request):
    admission_year, admission_round = get_current_year_and_round()
    try:
        candidates = admission_round.candidatemodel_set.all()
        if request.user.position not in [1, 2]:
            evaluations = admission_round.candidateevaluationmodel_set.all()
        else:
            evaluations = admission_round.candidateevaluationmodel_set.filter(
                evaluator=request.user)
    except (ObjectDoesNotExist, AttributeError):
        return render(request, 'mainapp/main_dashboard.html')
    context = {
        'candidates': candidates,
        'evaluations': evaluations,
        'admission_round': admission_round.round_number
    }
    return render(request, 'mainapp/main_dashboard.html', context)


@login_required(login_url='login')
@check_permissions(allowed_pos=[COMMITTEE_MEMBER, COMMITTEE_CHAIR])
def candidateEvaluateView(request, uuid):
    evaluator = request.user
    candidate = get_object_or_404(CandidateModel, candidate_id=uuid)
    evaluation = get_object_or_404(
        CandidateEvaluationModel,
        evaluator=evaluator,
        candidate=candidate)
    application_formset = ApplicationFormset(instance=evaluation)
    interview_formset = InterviewFormset(instance=evaluation)
    if request.method == "POST":
        application_formset = ApplicationFormset(
            request.POST, instance=evaluation)
        interview_formset = InterviewFormset(
            request.POST, instance=evaluation)
        if application_formset.is_valid() and interview_formset.is_valid():
            evaluation.evaluation_status = CandidateEvaluationModel.in_progress
            application_formset.save()
            interview_formset.save()
            evaluation.save()
            return redirect('dashboard')
    context = {
        'application_formset': application_formset,
        'interview_formset': interview_formset,
        'candidate': candidate}
    return render(request, 'mainapp/evaluate_candidate.html', context)


def queryset_to_dict(queryset, exclude=None):
    fields = queryset._meta.get_fields(include_hidden=False)
    print(fields)
    instance_dict = {}
    for field in fields:
        try:
            if field.name in exclude:
                continue
            key = field.verbose_name
            instance_dict[key] = getattr(queryset, field.name)
        except Exception:
            continue
    return instance_dict


@login_required(login_url='login')
@check_permissions(allowed_pos=[SECRETARY])
def approveEvalView(request, uuid):
    evaluation = get_object_or_404(
        CandidateEvaluationModel, evaluation_id=uuid)
    application_evaluation = get_object_or_404(
        ApplicationEvaluationModel, evaluation=evaluation)
    interview_evaluation = get_object_or_404(
        InterviewEvaluationModel, evaluation=evaluation)
    application_evaluation_dict = queryset_to_dict(
        application_evaluation,
        exclude=['evaluation', 'id'])
    if not interview_evaluation.skip_evaluation:
        interview_evaluation_dict = queryset_to_dict(
            interview_evaluation,
            exclude=['evaluation', 'id'])
    else:
        interview_evaluation_dict = None
    approve_form = ApprovementForm()
    evaluate_form = SecretaryEvaluationForm(instance=evaluation.candidate)
    if request.method == "POST":
        approve_form = ApprovementForm(request.POST)
        evaluate_form = SecretaryEvaluationForm(
            request.POST, instance=evaluation.candidate)
        if approve_form.is_valid() and evaluate_form.is_valid():
            evaluation.evaluation_status = approve_form.cleaned_data['status']
            evaluate_form.save()
            evaluation.save()
            return redirect('dashboard')

    context = {
        'evaluate_form': evaluate_form,
        'approve_form': approve_form,
        'application_evaluation_dict': application_evaluation_dict,
        'interview_evaluation_dict': interview_evaluation_dict,
        'evaluation': evaluation}

    return render(request, 'mainapp/approve_evaluation.html', context)


@login_required(login_url='login')
@check_permissions(allowed_pos=[ADMISSION_DEPARTMENT])
def createCandidateView(request, candidate_id=None):
    if candidate_id is not None:
        candidate = get_object_or_404(
            CandidateModel, candidate_id=candidate_id)
        form = AddCandidateForm(instance=candidate)
        testing_formset = TestingFormset(instance=candidate)
        education_formset = EducationFormset(instance=candidate)
    else:
        form = AddCandidateForm()
        testing_formset = TestingFormset()
        education_formset = EducationFormset()
    if request.method == "POST":
        form = AddCandidateForm(request.POST, request.FILES)
        testing_formset = TestingFormset(request.POST, instance=candidate)
        education_formset = EducationFormset(request.POST, instance=candidate)
        if form.is_valid() and testing_formset.is_valid() and   \
                education_formset.is_valid():
            admission_year = get_object_or_404(AdmissionYearModel, active=True)
            form.save()
            testing_formset.save()
            education_formset.save()
            candidate.student_list = admission_year.current_candidates
            return redirect('dashboard')
    context = {
        'form': form,
        'testing_formset': testing_formset,
        'education_formset': education_formset}
    return render(request, 'mainapp/create_candidate.html', context)


@login_required(login_url='login')
def infoView(request):
    publications = InformationModel.objects.all()
    context = {'publications': publications}
    return render(request, 'mainapp/info.html', context)


@login_required(login_url='login')
def contactsView(request):
    admission_year = get_object_or_404(AdmissionYearModel, active=True)
    staff_list = admission_year.get_staff_list()
    dept_members = staff_list.filter(position=ADMISSION_DEPARTMENT)
    committee_members = staff_list.filter(position=COMMITTEE_MEMBER)
    chairs = staff_list.filter(position=COMMITTEE_CHAIR)
    secretaries = staff_list.filter(position=SECRETARY)
    context = {
        'dept_members': dept_members,
        'committee': committee_members,
        'chairs': chairs,
        'secretaries': secretaries,
    }
    return render(request, 'mainapp/contacts.html', context)


@login_required(login_url='login')
def personalView(request):
    if request.method == 'POST':
        print("\n\n\nPOST INIT")
        form = CustomPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            print("\n\n\nPass changed")
            update_session_auth_hash(request, user)  # Important!
            messages.success(
                request, 'Your password was successfully updated!')
            return redirect('personal')
        else:
            print("\n\n\nError")
            messages.error(request, 'Please correct the error below.')
    else:
        form = CustomPasswordChangeForm(request.user)
    context = {
        'position': POSITIONS[request.user.position],
        'form': form,
        'user': request.user
    }
    return render(request, 'mainapp/personal.html', context)


@login_required(login_url='login')
def profileView(request, uuid):
    if uuid.replace(" ", "") == str(request.user.staff_id):
        return redirect('personal')
    context = {
        'position': POSITIONS[request.user.position],
        'user': get_object_or_404(get_user_model(), staff_id=uuid)
    }
    return render(request, 'mainapp/personal.html', context)


@auth_check
def loginView(request):
    context = {}
    if request.method == "POST":
        email = request.POST.get('email')
        password = request.POST.get('password')

        user = authenticate(request, email=email, password=password)

        if user is not None:
            login(request, user)
            return redirect('dashboard')

        else:
            messages.error(request, "Please check username and password")
            return render(request, 'mainapp/login.html', context)

    return render(request, 'mainapp/login.html', context)


def logoutView(request):
    logout(request)
    return redirect('login')


@login_required(login_url='login')
@check_permissions(allowed_pos=[COMMITTEE_CHAIR])
def ChairView(request):
    admission_year, admission_round = get_current_year_and_round()
    try:
        candidates_waiting_list = admission_year.current_candidates.    \
            candidatemodel_set.all()
        candidates_accepted = admission_round.  \
            accepted_candidates_list.candidatemodel_set.all()
        candidates = candidates_waiting_list | candidates_accepted
        evaluated_candidates_count = len(
            candidates.filter(evaluation_finished=True))
        non_evaluated_candidates_count = len(
            candidates.exclude(evaluation_finished=True))
    except (ObjectDoesNotExist, AttributeError):
        return redirect('dashboard')
    form = AdmissionRoundForm(instance=admission_round)
    context = {
        'form': form,
        'candidates': candidates.order_by("-total_score"),
        'total_candidates': len(candidates),
        'evaluated_candidates_count': evaluated_candidates_count,
        'non_evaluated_candidates_count': non_evaluated_candidates_count
    }
    if request.method == "POST":
        form = AdmissionRoundForm(request.POST, instance=admission_round)
        if non_evaluated_candidates_count:
            messages.error(
                request, 'Evaluation of candidates is not finished yet')
            return render(request, 'mainapp/chair_template.html', context)
        if form.is_valid():
            form.save()
            all_candidates = admission_year.    \
                current_candidates.candidatemodel_set.all()
            threshold = form.cleaned_data["threshold"]
            compose_lists(
                threshold, all_candidates, admission_year, admission_round)
            return redirect('dashboard')
    return render(request, 'mainapp/chair_template.html', context)


def clear_list(candidates, admission_year):
    for candidate in candidates:
        candidate.student_list = admission_year.current_candidates
        candidate.save()


def compose_lists(threshold, candidates, admission_year, admission_round):
    accepted_list = admission_round.    \
        accepted_candidates_list.candidatemodel_set.all()
    rejected_list = admission_round.    \
        rejected_candidates_list.candidatemodel_set.all()
    if accepted_list or rejected_list:
        print("\n\n\n clearing lists")
        clear_list(accepted_list, admission_year)
        clear_list(rejected_list, admission_year)
    for candidate in candidates:
        if candidate.total_score >= threshold:
            candidate.student_list = admission_round.accepted_candidates_list
        else:
            continue
        candidate.save()


@login_required(login_url='login')
@check_permissions(allowed_pos=[COMMITTEE_CHAIR])
def SecretaryView(request):
    admission_year, admission_round = get_current_year_and_round()
    try:
        waiting_list_candidates = admission_year.   \
            current_candidates.candidatemodel_set.all()
        recommended_list = \
            admission_round.accepted_candidates_list.candidatemodel_set.all()
        rejected_list_candidates = admission_round. \
            rejected_candidates_list.candidatemodel_set.all()
    except Exception as e:
        print(e)
    context = {
        'recommended_list_candidates': recommended_list,
        'waiting_list_candidates': waiting_list_candidates,
        'rejected_list_candidates': rejected_list_candidates,
    }
    return render(request, 'mainapp/secretary_template.html', context)


"""
        API views
"""


class CandidateDetail(generics.RetrieveUpdateDestroyAPIView):
    """
    Retrieve, update or delete a Candidate instance.
    """
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    lookup_url_kwarg = 'candidate_id'
    lookup_field = 'candidate_id'
    serializer_class = CandidateSerializer

    def get_queryset(self):
        return CandidateModel.objects.all()


class CandidatesList(generics.ListCreateAPIView):
    """
        Returns all candidate objects that are stored in the database
    """
    permission_classes = [permissions.IsAuthenticated]
    queryset = CandidateModel.objects.all()
    serializer_class = CandidateSerializer


"""
    RENDER EXCEL FILES AND SEND BACK AS A RESPONSE
"""


@login_required(login_url='login')
@check_permissions(allowed_pos=[SECRETARY])
def GetApplicationEvaluationAsExcelView(request, evaluation_id):
    """
        Render application evaluation into excel and send back to client
    """
    DEGREE = {
        0: 'BSc',
        1: 'MSc',
        2: 'PhD',
    }
    evaluation = CandidateEvaluationModel.objects.get(
        evaluation_id=evaluation_id)
    candidate = evaluation.candidate
    evaluator = evaluation.evaluator
    # set response type to excel file
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] =   \
        'attachment; filename="application_evaluation_{0}_{1}.xls"' \
        .format(candidate.first_name, candidate.last_name)
    # open excel template
    url = 'static/xlsx/application_template.xlsx'
    wb = load_workbook(url)
    sheets = wb.sheetnames
    work_sheet = wb[sheets[0]]
    # candidate general info
    work_sheet['B4'] = candidate.first_name
    work_sheet['C4'] = candidate.last_name
    work_sheet['D4'] = candidate.candidate_id
    work_sheet['F4'] = DEGREE[candidate.applying_degree]
    # evaluator general info
    work_sheet['B5'] = evaluator.first_name
    work_sheet['C5'] = evaluator.last_name
    work_sheet['D5'] = str(evaluator.staff_id)
    # candidate education information
    for instance in candidate.education_info.all():
        row_list = [instance.start_date, instance.end_date,
                    instance.grad_date, instance.degree_type,
                    instance.institution,
                    instance.study_field, instance.gpa]
        row_num = 8
        for col_num, value in enumerate(row_list):
            work_sheet.cell(row=row_num, column=col_num + 1).value = value
            print(value)
            print(str(row_num) + " " + str(col_num + 1))
            print("\n")
        row_num = row_num + 1
    # candidate testing information
    work_sheet['B12'] = candidate.testing_info.ielts
    work_sheet['E12'] = candidate.testing_info.gre
    work_sheet['B13'] = candidate.testing_info.toefl
    # evaluated by secretary
    work_sheet['F16'] = candidate.gpa
    work_sheet['F17'] = candidate.school_rating
    work_sheet['F18'] = candidate.research_experience
    # evaluated by evaluator
    work_sheet['F19'] = evaluation.application_evaluation.relevancy
    work_sheet['F20'] = evaluation.application_evaluation.statement_of_purpose
    work_sheet['F21'] = evaluation.application_evaluation.recommendation_1
    work_sheet['F22'] = evaluation.application_evaluation.recommendation_2
    work_sheet['F23'] = evaluation.application_evaluation.relevant_degrees
    work_sheet['B24'] = evaluation.application_evaluation.evaluation_comment
    wb.save(response)
    return response


@login_required(login_url='login')
@check_permissions(allowed_pos=[SECRETARY])
def GetInterviewEvaluationAsExcellView(request, evaluation_id):
    """
        Render interview evaluation into excel and send back to client
    """
    degree = {
        0: 'BSc',
        1: 'MSc',
        2: 'PhD',
    }
    evaluation = CandidateEvaluationModel.objects.get(
        evaluation_id=evaluation_id)
    candidate = evaluation.candidate
    evaluator = evaluation.evaluator
    # set response type to excell file
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] =   \
        'attachment; filename="interview_evaluation_{0}_{1}.xls"' \
        .format(candidate.first_name, candidate.last_name)
    # open excel template
    url = 'static/xlsx/interview_template.xlsx'
    wb = load_workbook(url)
    sheets = wb.sheetnames
    work_sheet = wb[sheets[0]]
    # candidate general info
    work_sheet['B4'] = candidate.first_name
    work_sheet['C4'] = candidate.last_name
    work_sheet['D4'] = candidate.candidate_id
    work_sheet['F4'] = degree[candidate.applying_degree]
    # evaluator general info
    work_sheet['B5'] = evaluator.first_name
    work_sheet['C5'] = evaluator.last_name
    work_sheet['D5'] = str(evaluator.staff_id)
    # evaluated by evaluator
    work_sheet['F8'] = evaluation.intervew_evaluation.work_experience_goals
    work_sheet['F9'] = evaluation.intervew_evaluation.  \
        research_interest_and_motivation
    work_sheet['F10'] = evaluation.intervew_evaluation.understanding_of_major
    work_sheet['F11'] = evaluation.intervew_evaluation.community_involvement
    work_sheet['F12'] = evaluation.intervew_evaluation.interpersonal_skills
    work_sheet['F13'] = evaluation.intervew_evaluation.english_level
    work_sheet['B14'] = evaluation.intervew_evaluation.interview_comment
    wb.save(response)
    return response
